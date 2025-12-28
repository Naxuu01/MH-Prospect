"""
Module d'export des prospects en CSV, Excel et PDF.
"""
import csv
import json
import sqlite3
import logging
from datetime import datetime
from typing import List, Dict, Any, Optional
from pathlib import Path

logger = logging.getLogger(__name__)


def exporter_csv(db_path: str, output_file: Optional[str] = None) -> str:
    """
    Exporte les prospects en CSV avec formatage amélioré.
    
    Args:
        db_path: Chemin vers la base de données
        output_file: Fichier de sortie (par défaut: prospects_export_YYYYMMDD.csv)
    
    Returns:
        Chemin du fichier créé
    """
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"prospects_export_{timestamp}.csv"
    
    try:
        conn = sqlite3.connect(db_path, timeout=10.0)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Vérifier les colonnes disponibles
        cursor.execute("PRAGMA table_info(prospects)")
        available_columns = [row[1] for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM prospects ORDER BY COALESCE(score, 0) DESC, date_traitement DESC")
        prospects = cursor.fetchall()
        
        if not prospects:
            logger.warning("Aucun prospect à exporter.")
            conn.close()
            return output_file
        
        # Colonnes prioritaires à exporter (dans l'ordre)
        priority_columns = [
            'id', 'nom_entreprise', 'site_web', 'telephone', 'email', 
            'email_status', 'linkedin_entreprise', 'score',
            'point_specifique', 'raison_choix', 'proposition_service',
            'message_personnalise', 'technologies', 'taille_entreprise',
            'industrie', 'note_google', 'nb_avis', 'template_utilise',
            'date_ajout', 'date_traitement', 'statut'
        ]
        
        # Filtrer pour ne garder que les colonnes existantes
        columns = [col for col in priority_columns if col in available_columns]
        # Ajouter les autres colonnes disponibles
        other_columns = [col for col in available_columns if col not in columns]
        columns.extend(other_columns)
        
        with open(output_file, 'w', newline='', encoding='utf-8') as f:
            writer = csv.DictWriter(f, fieldnames=columns, extrasaction='ignore')
            writer.writeheader()
            
            for prospect in prospects:
                row = {}
                for col in columns:
                    value = prospect.get(col, '')
                    # Nettoyer les valeurs pour CSV
                    if value is None:
                        value = ''
                    elif isinstance(value, str):
                        # Remplacer les retours à la ligne par des espaces
                        value = value.replace('\n', ' ').replace('\r', ' ')
                    row[col] = value
                writer.writerow(row)
        
        conn.close()
        logger.info(f"✅ {len(prospects)} prospects exportés vers {output_file}")
        return output_file
        
    except sqlite3.Error as e:
        logger.error(f"❌ Erreur DB lors de l'export CSV: {e}")
        raise
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'export CSV: {e}", exc_info=True)
        raise


def exporter_excel(db_path: str, output_file: Optional[str] = None) -> str:
    """
    Exporte les prospects en Excel avec formatage amélioré.
    
    Args:
        db_path: Chemin vers la base de données
        output_file: Fichier de sortie (par défaut: prospects_export_YYYYMMDD.xlsx)
    
    Returns:
        Chemin du fichier créé
    """
    try:
        import pandas as pd
        from openpyxl.styles import Font, PatternFill, Alignment
        from openpyxl.utils import get_column_letter
    except ImportError:
        logger.warning("pandas/openpyxl non installé. Export en CSV à la place.")
        if output_file:
            output_file = output_file.replace('.xlsx', '.csv')
        return exporter_csv(db_path, output_file)
    
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"prospects_export_{timestamp}.xlsx"
    
    try:
        conn = sqlite3.connect(db_path, timeout=10.0)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Vérifier les colonnes disponibles
        cursor.execute("PRAGMA table_info(prospects)")
        available_columns = [row[1] for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM prospects ORDER BY COALESCE(score, 0) DESC, date_traitement DESC")
        prospects = cursor.fetchall()
        
        if not prospects:
            logger.warning("Aucun prospect à exporter.")
            conn.close()
            return output_file
        
        # Convertir en DataFrame
        data = []
        for row in prospects:
            row_dict = {}
            for col in available_columns:
                value = row.get(col, '')
                if value is None:
                    value = ''
                row_dict[col] = value
            data.append(row_dict)
        
        df = pd.DataFrame(data)
        
        # Réorganiser les colonnes par priorité
        priority_cols = [
            'id', 'nom_entreprise', 'site_web', 'telephone', 'email', 
            'email_status', 'score', 'linkedin_entreprise',
            'point_specifique', 'raison_choix', 'proposition_service',
            'message_personnalise', 'technologies', 'taille_entreprise',
            'industrie', 'note_google', 'nb_avis', 'template_utilise',
            'date_ajout', 'date_traitement', 'statut'
        ]
        
        existing_cols = [col for col in priority_cols if col in df.columns]
        other_cols = [col for col in df.columns if col not in priority_cols]
        df = df[existing_cols + other_cols]
        
        # Export Excel avec formatage avancé
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Prospects')
            
            worksheet = writer.sheets['Prospects']
            
            # Style de l'en-tête
            header_fill = PatternFill(start_color="667eea", end_color="667eea", fill_type="solid")
            header_font = Font(bold=True, color="FFFFFF")
            
            for cell in worksheet[1]:
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Largeur des colonnes automatique
            for idx, col in enumerate(df.columns, 1):
                col_letter = get_column_letter(idx)
                max_length = max(
                    df[col].astype(str).map(len).max() if len(df) > 0 else 0,
                    len(str(col))
                )
                worksheet.column_dimensions[col_letter].width = min(max_length + 2, 60)
            
            # Formatage conditionnel pour la colonne score
            if 'score' in df.columns:
                score_col_idx = df.columns.get_loc('score') + 1
                score_col_letter = get_column_letter(score_col_idx)
                
                for row_idx in range(2, len(df) + 2):
                    cell = worksheet[f"{score_col_letter}{row_idx}"]
                    try:
                        score_val = int(cell.value) if cell.value else 0
                        if score_val >= 80:
                            cell.fill = PatternFill(start_color="d1fae5", end_color="d1fae5", fill_type="solid")
                        elif score_val >= 60:
                            cell.fill = PatternFill(start_color="dbeafe", end_color="dbeafe", fill_type="solid")
                        elif score_val >= 40:
                            cell.fill = PatternFill(start_color="fef3c7", end_color="fef3c7", fill_type="solid")
                        else:
                            cell.fill = PatternFill(start_color="fee2e2", end_color="fee2e2", fill_type="solid")
                    except (ValueError, TypeError):
                        pass
        
        conn.close()
        logger.info(f"✅ {len(prospects)} prospects exportés vers {output_file}")
        return output_file
        
    except sqlite3.Error as e:
        logger.error(f"❌ Erreur DB lors de l'export Excel: {e}")
        raise
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'export Excel: {e}", exc_info=True)
        raise


def exporter_pdf(db_path: str, output_file: Optional[str] = None) -> str:
    """
    Exporte les prospects en PDF (utilise reportlab si disponible).
    
    Args:
        db_path: Chemin vers la base de données
        output_file: Fichier de sortie (par défaut: prospects_export_YYYYMMDD.pdf)
    
    Returns:
        Chemin du fichier créé
    """
    try:
        from reportlab.lib.pagesizes import A4
        from reportlab.lib import colors
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    except ImportError:
        logger.error("reportlab non installé. Installez-le avec: pip install reportlab")
        raise ImportError("reportlab est requis pour l'export PDF. Installez-le avec: pip install reportlab")
    
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"prospects_export_{timestamp}.pdf"
    
    try:
        conn = sqlite3.connect(db_path, timeout=10.0)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        # Vérifier les colonnes disponibles
        cursor.execute("PRAGMA table_info(prospects)")
        available_columns = [row[1] for row in cursor.fetchall()]
        
        cursor.execute("SELECT * FROM prospects ORDER BY COALESCE(score, 0) DESC, date_traitement DESC LIMIT 100")
        prospects = cursor.fetchall()
        
        if not prospects:
            logger.warning("Aucun prospect à exporter.")
            conn.close()
            return output_file
        
        # Créer le document PDF
        doc = SimpleDocTemplate(output_file, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Titre
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1a1a1a'),
            spaceAfter=30
        )
        story.append(Paragraph(f"Rapport de Prospects - {datetime.now().strftime('%d/%m/%Y %H:%M')}", title_style))
        story.append(Spacer(1, 0.5*cm))
        
        # Statistiques détaillées
        total = len(prospects)
        avec_email = sum(1 for p in prospects if p.get('email'))
        avec_telephone = sum(1 for p in prospects if p.get('telephone'))
        scores = [p.get('score', 0) for p in prospects if p.get('score')]
        score_moyen = sum(scores) / len(scores) if scores else 0
        
        summary_data = [
            ['Total Prospects', str(total)],
            ['Avec Email', f"{avec_email} ({avec_email/total*100:.1f}%)" if total > 0 else "0"],
            ['Avec Téléphone', f"{avec_telephone} ({avec_telephone/total*100:.1f}%)" if total > 0 else "0"],
            ['Score Moyen', f"{score_moyen:.1f}/100" if score_moyen > 0 else "N/A"]
        ]
        summary_table = Table(summary_data, colWidths=[8*cm, 9*cm])
        summary_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f5f7fa')])
        ]))
        story.append(summary_table)
        story.append(Spacer(1, 1*cm))
        
        # Tableau des prospects amélioré
        table_data = [['Score', 'Entreprise', 'Email', 'Téléphone', 'Technologies', 'Statut']]
        
        for prospect in prospects[:50]:  # Limiter à 50 pour le PDF
            score = prospect.get('score', 0) or 0
            nom = prospect.get('nom_entreprise', '')[:25]
            email = prospect.get('email', 'N/A')[:25]
            tel = prospect.get('telephone', 'N/A')[:15]
            techs = prospect.get('technologies', '')[:20] if prospect.get('technologies') else '-'
            statut = prospect.get('statut', 'nouveau')[:10]
            
            row = [str(score), nom, email, tel, techs, statut]
            table_data.append(row)
        
        prospect_table = Table(table_data, colWidths=[2*cm, 5*cm, 4.5*cm, 3*cm, 3.5*cm, 2*cm])
        prospect_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#667eea')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (0, 0), (0, -1), 'CENTER'),  # Score centré
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.white),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('FONTSIZE', (0, 1), (-1, -1), 8),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f8f9fa')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ]))
        
        story.append(Paragraph("Liste des Prospects", styles['Heading2']))
        story.append(Spacer(1, 0.3*cm))
        story.append(prospect_table)
        
        # Construire le PDF
        doc.build(story)
        
        conn.close()
        logger.info(f"✅ {len(prospects)} prospects exportés vers {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'export PDF: {e}")
        raise


def exporter_json(db_path: str, output_file: Optional[str] = None) -> str:
    """
    Exporte les prospects en JSON.
    
    Args:
        db_path: Chemin vers la base de données
        output_file: Fichier de sortie (par défaut: prospects_export_YYYYMMDD.json)
    
    Returns:
        Chemin du fichier créé
    """
    if not output_file:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = f"prospects_export_{timestamp}.json"
    
    try:
        conn = sqlite3.connect(db_path)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute("SELECT * FROM prospects ORDER BY score DESC, date_traitement DESC")
        prospects = cursor.fetchall()
        
        if not prospects:
            logger.warning("Aucun prospect à exporter.")
            conn.close()
            return output_file
        
        data = {
            "export_date": datetime.now().isoformat(),
            "total_prospects": len(prospects),
            "prospects": [dict(row) for row in prospects]
        }
        
        with open(output_file, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=2)
        
        conn.close()
        logger.info(f"✅ {len(prospects)} prospects exportés vers {output_file}")
        return output_file
        
    except Exception as e:
        logger.error(f"❌ Erreur lors de l'export JSON: {e}")
        raise

